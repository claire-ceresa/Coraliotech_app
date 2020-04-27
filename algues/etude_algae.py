import xlsxwriter as x
from openpyxl import load_workbook
from pandas import DataFrame
from Bio import Entrez
from objects.NCBI_Search import NCBI_Search
from objects.NCBI_Product import NCBI_Product

Entrez.email = "claire.ceresa@hotmail.fr"
#
# workbook = load_workbook(filename="algues\species.xlsx")
# #worksheet = workbook["Species"]
# worksheet = workbook["Species2"]
# datas = DataFrame(worksheet.values)
#
# request_beginning = " [Organism] AND complete[Title] AND cds[Title] AND biomol_mrna[PROP]"
# list_ids = []
#
# for data in datas.values:
#     request = data[0] + request_beginning
#     search = NCBI_Search(request=request)
#     list_ids = list_ids + search.get_list_ids()


request = "Chlorella [Organism] AND complete[Title] AND cds[Title] AND biomol_mrna[PROP] NOT unnamed [Title] NOT hypothetical [Title]"
search = NCBI_Search(request=request)
list_ids = search.get_list_ids()
print(len(list_ids))

workbook = x.Workbook("algae_datas.xlsx")
worksheet = workbook.add_worksheet("Datas")

for line, id in enumerate(list_ids):
    try:
        product = NCBI_Product(id=id)
        name = product.name
        taxo = product.species.taxonomy
        length = product.cds.length
        mw = product.molecular_weight
        species = product.species.species
        row = [species, id, name, length, mw]
        worksheet.write_row(line, 0, row)
    except Exception as e:
        row = [id, str(e)]
        worksheet.write_row(line, 0, row)
    else:
        print(str(line))

workbook.close()