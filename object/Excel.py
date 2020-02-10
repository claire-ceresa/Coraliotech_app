import openpyxl as o
from openpyxl.styles import *
from openpyxl.utils import *


class Excel:

    def __init__(self, organism, type):
        self.title = "files\\" + organism + "_" + type + ".xlsx"

        if not self.check_if_existed(self.title):
            workbook_type = o.load_workbook("files_type\\" + type + ".xlsx")
            workbook_type.save(self.title)

        self.workbook = o.load_workbook(self.title)

    def create_worksheet(self, name, datas):
        """create a worksheet in the workbook"""
        worksheet_type = self.workbook["Type"]
        worksheet = self.workbook.copy_worksheet(worksheet_type)
        worksheet.title = name
        worksheet["A1"] = name
        self.add_data(worksheet, datas)

    def close(self):
        """close the workbook"""
        self.workbook.save(self.title)
        self.workbook.close()

    def add_data(self, worksheet, datas):
        """adding the data to each cells"""
        for line in enumerate(datas):
            for column in enumerate(line[1].values()):

                try:
                    cell = worksheet.cell(line[0] + 4, column[0]+1)
                    cell.value = column[1]
                except Exception as e:
                    print(e)

    def check_if_existed(self, title):
        """check if Excel file already exists"""
        try:
            file = o.open(title)
            file.close()
            return True
        except FileNotFoundError:
            return False
